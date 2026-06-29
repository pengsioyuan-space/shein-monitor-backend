import os
import json
import time
import hmac
import hashlib
import requests
from datetime import datetime, timezone, timedelta

def get_time_range(days=2):
    now = datetime.now()
    start = now - timedelta(days=days)
    end = now

    return (
        start.strftime("%Y-%m-%d %H:%M:%S"),
        end.strftime("%Y-%m-%d %H:%M:%S")
    )
    
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("未安装 openpyxl，请先运行：pip install openpyxl")


# =========================
# 配置区
# =========================

DOMAIN = "https://openapi-erp.91miaoshou.com"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
KEY_FILE = os.path.join(SCRIPT_DIR, "miaoshou_key.txt")

ORDER_START_FROM, ORDER_START_TO = get_time_range(2)
# 如需手动覆盖接口时间，可填写下面两个值；留空时直接使用 ORDER_START_FROM / ORDER_START_TO。
# 妙手接口不允许 gmtCreateTo 晚于当前时间，脚本会自动截断到当前北京时间。
GMT_CREATE_FROM = ""
GMT_CREATE_TO = ""
API_CREATE_LOOKBACK_DAYS = 0
API_CREATE_FORWARD_DAYS = 0

# 可选筛选项。空值会自动不传。
PLATFORM = ""                      # 例如：SHEIN；不确定就留空
SHOP_IDS = []                       # 例如：["12345", "67890"]；不筛选就留空
APP_PACKAGE_STATUS = ""             # 例如：wait_seller_send；不筛选就留空
PLATFORM_ORDER_SNS = ""             # 多个订单号按文档用逗号分割；不筛选就留空
GMT_MODIFIED_FROM = ""              # 不筛选就留空
GMT_MODIFIED_TO = ""                # 不筛选就留空

# 妙手接口限制 pageSize 不能大于 100。
PAGE_SIZE = 100
# 妙手接口实际要求 page 不能小于 1，所以从 page=1 开始。
PAGE_START = 1
MAX_PAGES = 10000
REQUEST_SLEEP_SECONDS = 0.15

# 首次请求打印 body 和返回摘要；没有数据时把原始响应保存到 JSON，方便排查字段名/时间范围。
DEBUG_FIRST_PAGE = True
SAVE_RAW_RESPONSE_WHEN_EMPTY = True

# 妙手接口返回的订单创建时间是北京时间 UTC+8；无时区字符串统一按 UTC+8 解析。
CREATE_TIME_TZ_OFFSET_HOURS = 8
CREATE_TIME_TZ = timezone(timedelta(hours=CREATE_TIME_TZ_OFFSET_HOURS), name=f"UTC{CREATE_TIME_TZ_OFFSET_HOURS:+d}")

# 下单时间已经等同于接口 gmtCreateFrom/gmtCreateTo，所以不再用 orderInfo.gmtOrderStart 二次过滤。
LOCAL_FILTER_BY_ORDER_START = False

# 当前表头不再导出仓库信息，因此默认不调用线上物流列表兜底接口。
USE_ONLINE_PRODUCT_FALLBACK = False
ONLINE_PRODUCT_SLEEP_SECONDS = 0.08

PACKAGE_LIST_PATH = "/open/v1/order/package/fetch/search_package_list"
ONLINE_PRODUCT_PATH = "/open/v1/order/logistics_agent/manage/get_enable_online_product_list"

OUTPUT_XLSX = f"妙手包裹物流信息v6.3-下单时间-{ORDER_START_FROM[:10].replace('-', '.')}-{ORDER_START_TO[:10].replace('-', '.')}.xlsx"

FIELDNAMES = [
    "店铺",
    "订单编号",
    "订单创建时间",
    "已创建小时数",
    "妙手包裹号",
    "物流单号",
    "履约类型",
]

STATUS_MAP = {
    "unpaid": "未付款",
    "wait_confirmed": "待审核",
    "wait_seller_send": "待发货",
    "wait_receiver_confirm": "已发货",
    "finished": "已完成",
    "cancelled": "未发货退款/已关闭",
    "returned": "已发货退款/已退款",
    "refunding": "售后中",
}


# =========================
# 基础工具
# =========================

def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def join_unique(values, sep=" | "):
    out = []
    for v in values:
        s = safe_str(v)
        if s and s not in out:
            out.append(s)
    return sep.join(out)


def clean_secret(v):
    s = safe_str(v)
    for ch in ["\ufeff", "\u200b", "\u200c", "\u200d", "\u2060", "\u00a0"]:
        s = s.replace(ch, "")
    return "".join(s.strip().strip('"').strip("'").split())


def mask_key(v, left=4, right=4):
    s = clean_secret(v)
    if not s:
        return "空"
    if len(s) <= left + right:
        return "*" * len(s)
    return f"{s[:left]}...{s[-right:]}"


def read_key_file(path=KEY_FILE):
    if not os.path.exists(path):
        raise FileNotFoundError(f"找不到密钥文件：{path}")

    data = {}
    with open(path, "r", encoding="utf-8-sig") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            k, v = line.split("=", 1)
            data[k.strip().lower()] = clean_secret(v)

    app_key = data.get("appkey") or data.get("app_key") or data.get("x-app-key")
    app_secret = data.get("appsecret") or data.get("app_secret") or data.get("secret")

    if not app_key or not app_secret:
        raise ValueError("miaoshou_key.txt 需要包含 AppKey=... 和 AppSecret=...")

    return app_key, app_secret


def make_body_json(body):
    """
    签名用 bodyJson 必须与实际发送的 body 字符串一致。
    separators 去掉空格，ensure_ascii=False 保持中文不转义。
    """
    return json.dumps(body or {}, ensure_ascii=False, separators=(",", ":"))


def make_signature(app_key, app_secret, path, timestamp, body_json):
    """
    文档规则：
    sign = HmacSHA256(appSecret, appSecret + path + timestamp + appKey + bodyJson + appSecret)
    输出：hex 小写
    """
    sign_source = f"{app_secret}{path}{timestamp}{app_key}{body_json}{app_secret}"
    return hmac.new(
        app_secret.encode("utf-8"),
        sign_source.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest().lower()


def call_miaoshou_api(app_key, app_secret, path, body, retries=3, debug=False):
    url = DOMAIN + path
    body_json = make_body_json(body)

    for i in range(retries):
        timestamp = str(int(time.time()))  # 秒级 Unix 时间戳
        sign = make_signature(app_key, app_secret, path, timestamp, body_json)
        headers = {
            "Content-Type": "application/json",
            "x-app-key": app_key,
            "x-timestamp": timestamp,
            "x-sign": sign,
        }

        if debug:
            print("请求 URL：", url)
            print("签名 path：", path)
            print("bodyJson：", body_json)
            print("x-app-key：", mask_key(app_key))
            print("x-timestamp：", timestamp)
            print("x-sign：", sign)

        try:
            resp = requests.post(url, headers=headers, data=body_json.encode("utf-8"), timeout=45)
            try:
                data = resp.json()
            except Exception:
                data = {"http_status": resp.status_code, "ok": False, "text": resp.text}
            return data
        except Exception as e:
            if i == retries - 1:
                return {"result": "fail", "code": "REQUEST_ERROR", "message": str(e), "data": []}
            time.sleep(0.5 * (i + 1))

    return {"result": "fail", "code": "UNKNOWN", "message": "未知错误", "data": []}


# =========================
# JSON 提取工具
# =========================

def first_existing(d, keys):
    if not isinstance(d, dict):
        return ""

    for key in keys:
        if key in d and d.get(key) not in [None, ""]:
            return d.get(key)

    lower = {str(k).lower(): k for k in d.keys()}
    for key in keys:
        real = lower.get(str(key).lower())
        if real and d.get(real) not in [None, ""]:
            return d.get(real)

    return ""


def collect_values_by_keys(obj, keys):
    out = []
    target = {str(k).lower() for k in keys}

    if isinstance(obj, dict):
        for k, v in obj.items():
            if str(k).lower() in target and v not in [None, ""]:
                out.append(v)
            if isinstance(v, (dict, list)):
                out.extend(collect_values_by_keys(v, keys))
    elif isinstance(obj, list):
        for item in obj:
            out.extend(collect_values_by_keys(item, keys))

    return out


def first_value_deep(obj, keys):
    for v in collect_values_by_keys(obj, keys):
        s = safe_str(v)
        if s:
            return s
    return ""


def find_first_list(obj, preferred_keys=None):
    preferred_keys = preferred_keys or []
    keys = preferred_keys + [
        "list", "records", "rows", "items", "dataList", "pageList",
        "orderPackageList", "packageList", "packages", "opOrderPackageList", "opOrderPackages",
    ]

    if isinstance(obj, list):
        return [x for x in obj if isinstance(x, dict)]

    if not isinstance(obj, dict):
        return []

    for key in keys:
        v = obj.get(key)
        if isinstance(v, list):
            return [x for x in v if isinstance(x, dict)]

    # 常见返回：{"data": {"list": [...]}}
    data = obj.get("data")
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for key in keys:
            v = data.get(key)
            if isinstance(v, list):
                return [x for x in v if isinstance(x, dict)]

    # 再做浅层兜底
    for v in obj.values():
        if isinstance(v, dict):
            for key in keys:
                vv = v.get(key)
                if isinstance(vv, list):
                    return [x for x in vv if isinstance(x, dict)]

    return []


def parse_time(value):
    if value in [None, ""]:
        return None
    if isinstance(value, (int, float)):
        try:
            n = float(value)
            if n > 10_000_000_000:
                n = n / 1000
            return datetime.fromtimestamp(n, tz=timezone.utc).astimezone(CREATE_TIME_TZ)
        except Exception:
            return None

    text = safe_str(value)
    if not text:
        return None

    if text.isdigit():
        try:
            n = float(text)
            if n > 10_000_000_000:
                n = n / 1000
            return datetime.fromtimestamp(n, tz=timezone.utc).astimezone(CREATE_TIME_TZ)
        except Exception:
            pass

    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%dT%H:%M:%SZ",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(text, fmt)
            if text.endswith("Z"):
                dt = dt.replace(tzinfo=timezone.utc)
            elif dt.tzinfo is None:
                dt = dt.replace(tzinfo=CREATE_TIME_TZ)
            return dt.astimezone(CREATE_TIME_TZ)
        except Exception:
            pass

    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=CREATE_TIME_TZ)
        return dt.astimezone(CREATE_TIME_TZ)
    except Exception:
        return None


def format_time(value):
    dt = parse_time(value)
    if not dt:
        return safe_str(value)
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def shift_time_text(value, days):
    dt = parse_time(value)
    if not dt:
        return safe_str(value)
    return (dt + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")


def min_time_text(a, b):
    """返回两个时间字符串中较早的一个；解析失败时尽量保留可用值。"""
    da = parse_time(a)
    db = parse_time(b)
    if da and db:
        return (da if da <= db else db).strftime("%Y-%m-%d %H:%M:%S")
    return safe_str(a) or safe_str(b)


def get_now_text():
    return datetime.now(CREATE_TIME_TZ).strftime("%Y-%m-%d %H:%M:%S")


def get_api_create_range():
    """
    用户已确认：后台“下单时间”就是妙手创建时间。
    因此接口直接传 ORDER_START_FROM / ORDER_START_TO 到 gmtCreateFrom / gmtCreateTo。

    重要：妙手接口不允许 gmtCreateTo 晚于当前时间。
    所以如果 ORDER_START_TO 是当天 23:59:59，但当前还没到 23:59:59，
    接口结束时间会自动截断到当前北京时间，避免报错。
    """
    api_from = safe_str(GMT_CREATE_FROM) or safe_str(ORDER_START_FROM)
    raw_api_to = safe_str(GMT_CREATE_TO) or safe_str(ORDER_START_TO)
    now_text = get_now_text()
    api_to = min_time_text(raw_api_to, now_text)
    return api_from, api_to


def elapsed_hours(value, now_dt):
    dt = parse_time(value)
    if not dt:
        return ""
    return round((now_dt - dt).total_seconds() / 3600, 2)


# =========================
# 字段提取
# =========================

def extract_shop_name(pkg):
    # 优先取顶层 shopName；不要被 shopId 抢先匹配。
    return first_existing(pkg, ["shopName", "shopNick", "storeName", "shopCnName", "shop_name", "店铺"]) or first_existing(pkg, ["shopId", "shopIds"])


def extract_order_no(pkg):
    return first_value_deep(pkg, [
        "platformOrderSn", "platformOrderSns", "platformOrderNo", "platformOrderNos",
        "platformOrderId", "orderSn", "orderNo", "orderId", "sourceOrderNo",
        "buyerOrderNo", "tradeNo",
    ])


def extract_status(pkg):
    # 用户指定优先用 orderInfo.appOrderStatusText；没有时再用包裹状态文本/状态码。
    order_info = pkg.get("orderInfo") if isinstance(pkg, dict) else {}
    raw = first_existing(order_info, ["appOrderStatusText"]) or first_existing(pkg, ["appPackageStatusText"])
    if raw:
        return safe_str(raw)

    code = first_existing(order_info, ["appOrderStatus"]) or first_existing(pkg, ["appPackageStatus"])
    return STATUS_MAP.get(safe_str(code), safe_str(code))


def extract_create_time(pkg):
    # 用户已确认：后台下单时间就是妙手创建时间。
    # 因此优先取包裹本身的 gmtCreate；如果接口没有返回，再兜底取 orderInfo.gmtOrderStart。
    order_info = pkg.get("orderInfo") if isinstance(pkg, dict) else {}
    return first_existing(pkg, ["gmtCreate", "gmtCreated", "gmtCreateTime", "createTime", "createdTime", "packageCreateTime"]) or first_existing(order_info, ["gmtOrderStart"]) or first_value_deep(pkg, [
        "gmtCreate", "gmtCreated", "gmtCreateTime", "createTime", "createdTime",
        "packageCreateTime", "gmtOrderStart", "orderCreateTime", "platformOrderCreateTime",
    ])


def extract_package_id(pkg):
    return first_value_deep(pkg, [
        "opOrderPackageId", "orderPackageId", "packageId", "opPackageId", "id",
    ])


def extract_logistics_auth_id(pkg):
    return first_value_deep(pkg, [
        "logisticsAgentAuthId", "agentAuthId", "logisticsAuthId",
    ])


def extract_agent_code(pkg):
    return first_value_deep(pkg, [
        "agentCode", "logisticsAgentCode", "logisticsCode", "carrierCode",
    ])


def extract_warehouse_code(pkg):
    return first_value_deep(pkg, [
        "warehouseId", "warehouseCode", "warehouseNo", "warehouseSn",
        "deliveryWarehouseId", "shippingWarehouseId", "sendWarehouseId",
        "outWarehouseId", "stockWarehouseId", "warehouseAddressCode",
    ])


def extract_warehouse_name(pkg):
    return first_value_deep(pkg, [
        "warehouseName", "warehouse_name", "deliveryWarehouseName", "shippingWarehouseName",
        "sendWarehouseName", "outWarehouseName", "stockWarehouseName",
    ])


def extract_app_package_no(pkg):
    # “妙手包裹号”明确取包裹顶层 appPackageNo，例如：MS20260615212429038。
    # 只有顶层没有时，才兼容旧字段 packageNo / opPackageNo。
    v = first_existing(pkg, ["appPackageNo"])
    if safe_str(v):
        return safe_str(v)
    return first_existing(pkg, ["packageNo", "opPackageNo"])


def extract_logistics_field(pkg, field_names):
    """
    优先取 logisticsAgentProductInfo，其次取 opOrderPackageToPlatformLastMile，再全局深搜。
    """
    for block_name in ["logisticsAgentProductInfo", "opOrderPackageToPlatformLastMile"]:
        block = pkg.get(block_name) if isinstance(pkg, dict) else {}
        v = first_existing(block, field_names)
        if safe_str(v):
            return safe_str(v)
    return first_value_deep(pkg, field_names)


def extract_logistics_product_code(pkg):
    return extract_logistics_field(pkg, ["productCode", "logisticsProductCode", "logisticsAgentProductCode"])


def extract_logistics_product_name(pkg):
    return extract_logistics_field(pkg, ["productName", "combineProductName", "logisticsProductName", "logisticsAgentProductName"])


def extract_logistics_no(pkg):
    # “物流单号”明确优先取包裹顶层 logisticsNo，
    # 因为 logisticsAgentProductInfo / opOrderPackageToPlatformLastMile 里的 logisticsNo 可能为空。
    # 例如：顶层 logisticsNo = 9200190352018903941978。
    v = first_existing(pkg, ["logisticsNo"])
    if safe_str(v):
        return safe_str(v)
    return extract_logistics_field(pkg, ["logisticsNo", "trackingNo", "waybillNo", "platformPackageNo"])


def extract_logistics_company(pkg):
    return extract_logistics_field(pkg, ["logisticsCompany", "carrierName", "companyName", "logisticsName"])


def extract_fulfillment_type(pkg):
    raw = first_existing(pkg, ["fulfillmentType"])
    mapping = {
        "platformLogisticsFulfillment": "平台物流履约",
        "sellerFulfillment": "卖家自发货",
    }
    return mapping.get(safe_str(raw), safe_str(raw))



def extract_warehouse_from_online_products(result):
    """
    从“有效线上物流列表”结果里尽量提取仓库编码/名称。
    若多个候选，返回拼接结果。
    """
    items = find_first_list(result, [
        "onlineProductList", "enableOnlineProductList", "productList", "logisticsProductList",
        "warehouseList", "availableWarehouseList",
    ])
    codes = []
    names = []

    for item in items:
        c = extract_warehouse_code(item)
        n = extract_warehouse_name(item)
        if c:
            codes.append(c)
        if n:
            names.append(n)

    # 有些返回结构不是列表，直接在 result 里深搜。
    if not codes:
        c = extract_warehouse_code(result)
        if c:
            codes.append(c)
    if not names:
        n = extract_warehouse_name(result)
        if n:
            names.append(n)

    return join_unique(codes), join_unique(names)




def summarize_api_result(result):
    """打印返回结构摘要，避免控制台刷出太长 JSON。"""
    if not isinstance(result, dict):
        print("返回不是 dict：", type(result).__name__)
        return
    print("返回顶层字段：", list(result.keys()))
    data = result.get("data")
    if isinstance(data, dict):
        print("data 字段类型：dict，子字段：", list(data.keys()))
        for k, v in data.items():
            if isinstance(v, list):
                print(f"data.{k} 是列表，长度：{len(v)}")
            elif isinstance(v, dict):
                print(f"data.{k} 是对象，子字段：{list(v.keys())}")
    elif isinstance(data, list):
        print("data 字段类型：list，长度：", len(data))
    else:
        print("data 字段类型：", type(data).__name__)
    msg = safe_str(result.get("message") or result.get("msg"))
    if msg:
        print("接口 message：", msg)


def save_raw_response(result, filename="miaoshou_last_package_response.json"):
    path = os.path.join(SCRIPT_DIR, filename)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print("已保存原始返回：", path)
    except Exception as e:
        print("保存原始返回失败：", e)

# =========================
# 业务接口
# =========================

def build_package_search_body(page):
    api_create_from, api_create_to = get_api_create_range()
    body = {
        "page": page,
        # 妙手接口最大 pageSize=100，防止误改配置后接口报错。
        "pageSize": min(int(PAGE_SIZE), 100),
        "gmtCreateFrom": api_create_from,
        "gmtCreateTo": api_create_to,
    }

    if PLATFORM:
        body["platform"] = PLATFORM
    if SHOP_IDS:
        body["shopIds"] = [safe_str(x) for x in SHOP_IDS if safe_str(x)]
    if APP_PACKAGE_STATUS:
        body["appPackageStatus"] = APP_PACKAGE_STATUS
    if PLATFORM_ORDER_SNS:
        body["platformOrderSns"] = PLATFORM_ORDER_SNS
    if GMT_MODIFIED_FROM:
        body["gmtModifiedFrom"] = GMT_MODIFIED_FROM
    if GMT_MODIFIED_TO:
        body["gmtModifiedTo"] = GMT_MODIFIED_TO

    return body


def fetch_package_page(app_key, app_secret, page):
    body = build_package_search_body(page)
    debug = DEBUG_FIRST_PAGE and page == PAGE_START
    result = call_miaoshou_api(app_key, app_secret, PACKAGE_LIST_PATH, body, debug=debug)
    return result


def fetch_all_packages(app_key, app_secret):
    all_packages = []

    for page in range(PAGE_START, PAGE_START + MAX_PAGES):
        api_page_size = min(int(PAGE_SIZE), 100)
        print(f"获取包裹列表 page={page}, pageSize={api_page_size}")
        result = fetch_package_page(app_key, app_secret, page)

        result_flag = safe_str(result.get("result") or result.get("code"))
        msg = safe_str(result.get("message") or result.get("msg"))
        if result_flag.lower() in ["fail", "error"]:
            print("包裹列表接口失败：", msg or result)
            break

        packages = find_first_list(result, [
            "orderPackageList", "packageList", "opOrderPackageList", "opOrderPackages", "list", "records",
        ])
        if page == PAGE_START:
            summarize_api_result(result)
        print(f"本页解析到 {len(packages)} 个包裹")
        if page == PAGE_START and not packages and SAVE_RAW_RESPONSE_WHEN_EMPTY:
            save_raw_response(result)
            print("提示：如果原始返回里 data/list 也是空，说明接口确实没有按当前条件查到包裹。")
            print("常见原因：1）当前时间范围内妙手没有创建包裹；2）需要加 platform 或 shopIds；3）包裹状态筛选 appPackageStatus 不匹配。")
        all_packages.extend(packages)

        if len(packages) < min(int(PAGE_SIZE), 100):
            break
        time.sleep(REQUEST_SLEEP_SECONDS)

    # 按包裹 id 或订单号去重
    unique = {}
    for pkg in all_packages:
        key = extract_app_package_no(pkg) or extract_package_id(pkg) or extract_logistics_no(pkg) or extract_order_no(pkg) or json.dumps(pkg, ensure_ascii=False, sort_keys=True)[:200]
        unique[key] = pkg

    out = list(unique.values())
    print("去重后包裹数：", len(out))
    return out


def fetch_online_product_warehouse(app_key, app_secret, pkg):
    op_order_package_id = extract_package_id(pkg)
    logistics_agent_auth_id = extract_logistics_auth_id(pkg)
    agent_code = extract_agent_code(pkg)
    warehouse_id = extract_warehouse_code(pkg)

    if not (op_order_package_id and logistics_agent_auth_id and agent_code):
        return "", ""

    body = {
        "opOrderPackageId": int(op_order_package_id) if safe_str(op_order_package_id).isdigit() else op_order_package_id,
        "logisticsAgentAuthId": int(logistics_agent_auth_id) if safe_str(logistics_agent_auth_id).isdigit() else logistics_agent_auth_id,
        "agentCode": agent_code,
    }
    if warehouse_id:
        body["warehouseId"] = warehouse_id

    result = call_miaoshou_api(app_key, app_secret, ONLINE_PRODUCT_PATH, body)
    time.sleep(ONLINE_PRODUCT_SLEEP_SECONDS)

    result_flag = safe_str(result.get("result") or result.get("code"))
    if result_flag.lower() in ["fail", "error"]:
        return "", ""

    return extract_warehouse_from_online_products(result)


def order_start_filter_reason(pkg):
    if not LOCAL_FILTER_BY_ORDER_START:
        return "keep"

    create_raw = extract_create_time(pkg)
    create_dt = parse_time(create_raw)
    start_dt = parse_time(ORDER_START_FROM)
    end_dt = parse_time(ORDER_START_TO)

    # 没有时间时先保留，避免误删；同时在统计里提示。
    if not create_dt:
        return "keep_missing_order_start"
    if not start_dt or not end_dt:
        return "keep_bad_config"
    if create_dt < start_dt:
        return "skip_before"
    if create_dt > end_dt:
        return "skip_after"
    return "keep"


def in_order_start_range(pkg):
    return order_start_filter_reason(pkg).startswith("keep")


def build_rows(app_key, app_secret, packages):
    # 获取当前北京时间 UTC+8，并用它和订单创建时间计算“已创建小时数”。
    now_dt = datetime.now(CREATE_TIME_TZ)
    rows = []
    filter_counter = {
        "keep": 0,
        "keep_missing_order_start": 0,
        "keep_bad_config": 0,
        "skip_before": 0,
        "skip_after": 0,
    }

    for idx, pkg in enumerate(packages, 1):
        reason = order_start_filter_reason(pkg)
        filter_counter[reason] = filter_counter.get(reason, 0) + 1
        if not reason.startswith("keep"):
            continue

        create_raw = extract_create_time(pkg)

        row = {
            "店铺": extract_shop_name(pkg),
            "订单编号": extract_order_no(pkg),
            "订单创建时间": format_time(create_raw),
            "已创建小时数": elapsed_hours(create_raw, now_dt),
            "妙手包裹号": extract_app_package_no(pkg),
            "物流单号": extract_logistics_no(pkg),
            "履约类型": extract_fulfillment_type(pkg),
        }
        rows.append(row)

        if idx % 50 == 0:
            print(f"已处理 {idx}/{len(packages)} 个候选包裹")

    print("本地 orderInfo.gmtOrderStart 过滤统计：", filter_counter)
    print("本地过滤后导出行数：", len(rows))
    if filter_counter.get("keep_missing_order_start"):
        print("提示：有包裹缺少 orderInfo.gmtOrderStart，脚本已保留，避免误删。")
    return rows

# =========================
# Excel 导出
# =========================

def export_xlsx(rows, filename=OUTPUT_XLSX):
    wb = Workbook()
    ws = wb.active
    ws.title = "包裹仓库信息"

    fill = PatternFill("solid", fgColor="D9F3EE")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, row in enumerate(rows, 2):
        for c, h in enumerate(FIELDNAMES, 1):
            value = row.get(h, "")
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if h == "已创建小时数" and value != "":
                cell.number_format = "0.00"
            else:
                cell.number_format = "@"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{max(ws.max_row, 1)}"

    widths = {
        "店铺": 22,
        "订单编号": 28,
        "订单创建时间": 22,
        "已创建小时数": 16,
        "妙手包裹号": 24,
        "物流单号": 28,
        "履约类型": 18,
    }
    for c, h in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 18)

    wb.save(filename)
    print("已生成 Excel：", filename)


# =========================
# 主程序
# =========================

def main():
    print("开始导出妙手包裹物流信息 v6.3 - 只走妙手")
    print("密钥文件：", KEY_FILE)
    print("接口域名：", DOMAIN)
    api_create_from, api_create_to = get_api_create_range()
    print("后台下单时间 = 妙手创建时间：", ORDER_START_FROM, "到", ORDER_START_TO)
    print("当前北京时间：", get_now_text())
    print("接口请求范围 gmtCreateFrom/gmtCreateTo：", api_create_from, "到", api_create_to)
    print("说明：如果结束时间晚于当前北京时间，脚本会自动截断到当前时间，避免接口报错。")
    print("分页从 page=", PAGE_START, "开始")
    print("本地 orderInfo.gmtOrderStart 二次过滤 LOCAL_FILTER_BY_ORDER_START=", LOCAL_FILTER_BY_ORDER_START)
    print("输出字段：", " | ".join(FIELDNAMES))

    app_key, app_secret = read_key_file(KEY_FILE)
    print("妙手 AppKey：", mask_key(app_key), "len=", len(app_key))
    print("妙手 AppSecret：", mask_key(app_secret), "len=", len(app_secret))

    packages = fetch_all_packages(app_key, app_secret)
    rows = build_rows(app_key, app_secret, packages)
    print("导出行数：", len(rows))
    export_xlsx(rows)
    print("完成。")


if __name__ == "__main__":
    main()
