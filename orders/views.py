from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from orders.models import Order

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None


def cors_json(data, status=200):
    response = JsonResponse(data, status=status, json_dumps_params={"ensure_ascii": False})
    response["Access-Control-Allow-Origin"] = "*"
    response["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    response["Access-Control-Allow-Headers"] = "Content-Type"
    return response


def cors_response(response):
    response["Access-Control-Allow-Origin"] = "*"
    response["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    response["Access-Control-Allow-Headers"] = "Content-Type"
    return response


def safe_float(value):
    try:
        if value in [None, ""]:
            return 0
        return float(value)
    except Exception:
        return 0


def order_to_dict(order):
    return {
        "order_no": getattr(order, "order_no", "") or "",
        "shop_name": getattr(order, "shop_name", "") or "",
        "region": getattr(order, "region", "") or "OTHER",
        "created_hours": safe_float(getattr(order, "created_hours", 0)),
        "logistics_no": getattr(order, "logistics_no", "") or "",
    }


@csrf_exempt
def dashboard(request):
    if request.method == "OPTIONS":
        return cors_json({})

    qs = Order.objects.all()

    total = qs.count()
    eu = qs.filter(region="EU").count()
    us = qs.filter(region="US").count()
    ca = qs.filter(region="CA").count()

    t12 = sum(1 for o in qs.only("created_hours") if safe_float(getattr(o, "created_hours", 0)) >= 12)
    t24 = sum(1 for o in qs.only("created_hours") if safe_float(getattr(o, "created_hours", 0)) >= 24)
    t36 = sum(1 for o in qs.only("created_hours") if safe_float(getattr(o, "created_hours", 0)) >= 36)
    t48 = sum(1 for o in qs.only("created_hours") if safe_float(getattr(o, "created_hours", 0)) >= 48)

    return cors_json({
        "total": total,
        "eu": eu,
        "us": us,
        "ca": ca,
        "other": max(total - eu - us - ca, 0),
        "t12": t12,
        "t24": t24,
        "t36": t36,
        "t48": t48,
    })


@csrf_exempt
def trend(request):
    if request.method == "OPTIONS":
        return cors_json({})

    qs = Order.objects.all()
    total = qs.count()
    buckets = {
        "12h+": 0,
        "24h+": 0,
        "36h+": 0,
        "48h+": 0,
    }

    for order in qs.only("created_hours"):
        h = safe_float(getattr(order, "created_hours", 0))
        if h >= 12:
            buckets["12h+"] += 1
        if h >= 24:
            buckets["24h+"] += 1
        if h >= 36:
            buckets["36h+"] += 1
        if h >= 48:
            buckets["48h+"] += 1

    region_counts = {"EU": 0, "US": 0, "CA": 0, "OTHER": 0}
    for row in qs.values("region").annotate(count=Count("id")):
        region = row.get("region") or "OTHER"
        if region not in region_counts:
            region = "OTHER"
        region_counts[region] += row.get("count", 0)

    return cors_json({
        "total": total,
        "risk": buckets,
        "region": region_counts,
    })


@csrf_exempt
def order_list(request):
    if request.method == "OPTIONS":
        return cors_json({})

    try:
        limit = int(request.GET.get("limit", 500))
    except Exception:
        limit = 500
    limit = max(1, min(limit, 2000))

    qs = Order.objects.all().order_by("-created_hours")[:limit]
    data = [order_to_dict(o) for o in qs]

    return cors_json({
        "count": Order.objects.count(),
        "data": data,
    })


@csrf_exempt
def export_ops_orders(request):
    if request.method == "OPTIONS":
        return cors_json({})

    if Workbook is None:
        return cors_json({"error": "openpyxl 未安装，请安装 openpyxl"}, status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "运营订单"

    headers = ["订单号", "店铺", "区域", "已创建小时数", "物流单号"]
    fill = PatternFill("solid", fgColor="D9F3EE")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index, order in enumerate(Order.objects.all().order_by("-created_hours"), 2):
        row = order_to_dict(order)
        values = [
            row["order_no"],
            row["shop_name"],
            row["region"],
            row["created_hours"],
            row["logistics_no"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if col == 4:
                cell.number_format = "0.00"
            else:
                cell.number_format = "@"

    widths = [28, 28, 12, 16, 32]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ops_orders.xlsx"'
    wb.save(response)
    return cors_response(response)
